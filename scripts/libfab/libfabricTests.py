import openpyxl
from pathlib import Path
import os
import subprocess
import re
import argparse
from datetime import datetime


def read_tests_excel(xlsx_file, list_tests, text_replace):
    '''
    @function: This method <read_tests_excel()> loads test details from xlsx excel file
               and returns list of test dictionaries
    @author:   venkatesh balagani
    @returns:  list of dictionaries <Ex: [{}, {},...]>
    '''

    # creating workbook obj for given xlsx file
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    # creating rows into lists <Ex: [[], [],...]>
    rowList = []
    for row in sheet.iter_rows(1, sheet.max_row):
        columnList = []
        for col in row:
            columnList.append(str(col.value).replace('xx.xxx.xxx.xx', text_replace))
        rowList.append(columnList)
        del columnList
    # print(rowList)
    
    # User Specific tests if mentioned in argument
    list_rows = []
    if len(list_tests) > 0:
        for test_case in list_tests:
            for row in rowList:
                if row[0] == test_case:
                    list_rows.append(row)
    else:
        list_rows = rowList
    # print(list_rows)
    
    '''
    creating list of test dictionaries with values in rows and with keys in ...
    variable names in first row <Ex: [{}, {},...]>
    '''
    row_index = 0
    list_test_dicts = []
    for row in list_rows:
        if row_index > 0 or (len(list_tests) > 0):
            list_test_dicts.append(dict(zip(rowList[0], row)))
        row_index = row_index + 1
    # print(list_test_dicts)
    return list_test_dicts

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def test_engine(list_test_dictionaries, stp_frst_flr):
    '''
    @function: This method picks up test by test from 'list_test_dictionaries' 
               and executes the same unto the last test
    @author:   venkatesh balagani
    @args:     list_test_dictionaries = List contains test dictionaries
               srvr_addr              = Server Address for Pingpong Test
               stp_frst_flr           = Stop on First Failure
    @returns:  Nothing
    '''

    # Log file definition <@Location: /tmp/libfabric_%H_%M_%d_%m_%Y.log>
    log_name = datetime.now().strftime('libfabric_%H_%M_%d_%m_%Y.log')
    fylPntr = open("/tmp/{}".format(log_name), "a")
    
    passCount = 0
    failCount = 0
    listFailedTests = []
    for test in list_test_dictionaries:  
        print("===========================================================")
        print(test['TestName'])

        match = 0
        if str(test['HW-specific']).upper() == "YES":
            hostname = os.popen("hostname")
            hostname = hostname.read()
            if re.search("vm", hostname.strip()):
                match = 1

        if match == 1:
            print("PASS | {}".format(test['message']))
            fylPntr.write("PASS | {} | {}\n".format(test['message'], test['TestName']))
            fylPntr.write("===============================================\n")
            fylPntr.write("\n{}".format(test['command']))
            fylPntr.write("\n------------------------------------------------\n")
            fylPntr.write("\n{}".format(output))
            fylPntr.write("\n------------------------------------------------\n")
            passCount = passCount + 1
        else:
            # if-server-instance-required
            if str(test['server']).upper() == "YES":
                start_server = subprocess.Popen(test['server-command'], shell=True)

            # command-exec
            output = os.popen(test['command'])
            output = output.read()

            # closing-server-instance-if-opened
            if str(test['server']).upper() == "YES":
                subprocess.Popen.kill(start_server)

            match = 0
            if str(test['match-sub']).upper() == "YES":
                if re.search(test['expected'], output.strip()):
                    match = 1
            else:
                if test['expected'] == output.strip():
                    match = 1

            if match == 1:
                print("PASS")
                fylPntr.write("PASS | {}\n".format(test['TestName']))
                fylPntr.write("===============================================\n")
                fylPntr.write("\n{}".format(test['command']))
                fylPntr.write("\n------------------------------------------------\n")
                fylPntr.write("\n{}".format(output))
                fylPntr.write("\n------------------------------------------------\n")
                passCount = passCount + 1

            else:
                print("FAIL")
                fylPntr.write("FAIL | {}\n".format(test['TestName']))
                fylPntr.write("===============================================\n")
                fylPntr.write("\n{}".format(test['command']))
                fylPntr.write("\n------------------------------------------------\n")
                fylPntr.write("\n{}".format(output))
                fylPntr.write("\n------------------------------------------------\n")
                failCount = failCount + 1
                listFailedTests.append(test['TestName'])
                
                if stp_frst_flr:
                    print("\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
                    print(test['TestName'])
                    print("Find full log: /tmp/{}".format(log_name))
                    raise Exception('Stopping on first test failure itself')
                    print("\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

    print("\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    print("Find full log: /tmp/{}".format(log_name))
    print("Total tests: ", len(list_test_dicts))
    print("PASS: {}".format(passCount))
    print("FAIL: {}".format(failCount))
    print("Failed Tests: ", listFailedTests)
    print("\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~    
# Command line Arguments <PARSER>
parser = argparse.ArgumentParser(description="Basic Arguments to run Libfabric")
parser.add_argument('-xl', action='store', default="libfabric.xlsx", dest='excelFile', help='Excel XLSX File')
parser.add_argument('-srvr_addr', action='store', default="127.0.0.1", dest='serverAddress', help='Server Address')
parser.add_argument('-tests', nargs='+', default=[], dest='testCasesList', help='Provide specifi Tests to run <Ex: [test1, test2,...], in default run all tests>')
parser.add_argument('-stop_on_error', action='store', default=False, dest='returnOnFirstFailure', help='Stop the script '
                                                                                                    'on first failure')
args = parser.parse_args()
list_of_TCs = list(args.testCasesList)
print('Excel File Path                     = {!r}'.format(args.excelFile))
print('Server Address                      = {!r}'.format(args.serverAddress))
print('Tests to Run                        = {!r}'.format(list_of_TCs))
print('Stop the script on First Failure    = {!r}'.format(args.returnOnFirstFailure))

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~    
# Calling Functions
list_test_dicts = read_tests_excel(args.excelFile, list_of_TCs, args.serverAddress)
test_engine(list_test_dicts, args.returnOnFirstFailure)
